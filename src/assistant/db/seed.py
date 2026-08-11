"""Build data/assets.db from Sample_Asset_Master.xlsx.

Idempotent: safe to re-run. Two responsibilities kept deliberately separate so a
reviewer can audit exactly what is real and what is not:

  1. Load the 21 source rows from the spreadsheet, unchanged  -> source='xlsx'
  2. Add the employee hierarchy and spare inventory the source
     file does not contain but the brief requires                -> source='synthetic'

Run `python -m assistant.db.seed --verify` to assert the 21 source rows in the
database still match the spreadsheet byte for byte.
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from assistant.config import get_settings

# Anchor for deriving `condition` from purchase date. Fixed rather than
# `date.today()` so seeding is reproducible: the same spreadsheet always
# produces the same database.
CONDITION_ANCHOR = date(2026, 1, 1)

EMAIL_DOMAIN = "xyztech.example"  # reserved TLD — never routes anywhere real

# --------------------------------------------------------------------------
# SYNTHETIC. The spreadsheet has no manager, department or e-mail column, but
# requirement 3 ("who is that employee's manager?") needs one. Declared as a
# literal table rather than generated so it is reviewable at a glance.
# name -> (department, base_location, manager_name)
# --------------------------------------------------------------------------
ORG_CHART: dict[str, tuple[str, str, str | None]] = {
    "Vikram Shah": ("IT Infrastructure", "Mumbai", None),          # Head of IT
    "Priya Singh": ("Engineering", "Pune", "Vikram Shah"),         # Engineering Manager
    "Anjali Desai": ("Operations", "Mumbai", "Vikram Shah"),       # Operations Manager
    "Sakshi Jain": ("Design", "Bangalore", "Vikram Shah"),         # Design Manager
    "Rahul Sharma": ("Engineering", "Pune", "Priya Singh"),
    "Amit Kumar": ("Engineering", "Bangalore", "Priya Singh"),
    "Arjun Rao": ("Engineering", "Bangalore", "Priya Singh"),
    "Deepak Sinha": ("Engineering", "Mumbai", "Priya Singh"),
    "Rohan Gupta": ("Operations", "Hyderabad", "Anjali Desai"),
    "Nitin Patil": ("Operations", "Hyderabad", "Anjali Desai"),
    "Karan Mehta": ("Operations", "Kolkata", "Anjali Desai"),
    "Sneha Verma": ("Operations", "Kolkata", "Anjali Desai"),
    "Neha Joshi": ("Design", "Chennai", "Sakshi Jain"),
}

# --------------------------------------------------------------------------
# SYNTHETIC. The spreadsheet has no status column and every row is assigned, so
# requirement 5 ("find an available laptop in Bangalore") would return nothing.
# This is unassigned stock plus two in-repair units, which give the
# recommendation tool something to correctly *exclude*.
# (code, name, category, location, purchase_date, status, condition)
# --------------------------------------------------------------------------
SPARE_INVENTORY: list[tuple[str, str, str, str, str, str, str]] = [
    ("AST1021", "Dell Latitude 7440", "Laptop", "Bangalore", "2025-06-10", "Available", "New"),
    ("AST1022", "MacBook Pro M3", "Laptop", "Bangalore", "2025-09-02", "Available", "New"),
    ("AST1023", "Lenovo ThinkPad X1", "Laptop", "Bangalore", "2024-03-18", "Available", "Good"),
    ("AST1024", "HP EliteBook 840", "Laptop", "Pune", "2025-01-20", "Available", "New"),
    ("AST1025", "Dell Latitude 7440", "Laptop", "Mumbai", "2024-11-05", "Available", "Good"),
    ("AST1026", "iPhone 15", "Mobile", "Bangalore", "2025-04-14", "Available", "New"),
    ("AST1027", "Dell Monitor 27", "Monitor", "Bangalore", "2024-08-22", "Available", "Good"),
    ("AST1028", "Dell Monitor 27", "Monitor", "Chennai", "2023-12-01", "Available", "Fair"),
    ("AST1029", "Canon Printer", "Printer", "Hyderabad", "2024-02-09", "Available", "Good"),
    ("AST1030", "Zebra TC22 Scanner", "Scanner", "Pune", "2025-02-28", "Available", "New"),
    ("AST1031", "MacBook Pro M3", "Laptop", "Delhi", "2023-05-16", "Available", "Fair"),
    ("AST1032", "iPhone 15", "Mobile", "Mumbai", "2024-06-30", "Available", "Good"),
    ("AST1033", "HP EliteBook 840", "Laptop", "Bangalore", "2023-03-12", "In Repair", "Poor"),
    ("AST1034", "HP LaserJet Pro", "Printer", "Kolkata", "2023-08-19", "In Repair", "Fair"),
]


def _email_for(name: str) -> str:
    return name.lower().replace(" ", ".") + "@" + EMAIL_DOMAIN


def _condition_for(purchase_date: date) -> str:
    """Deterministic age -> condition mapping (synthetic, documented)."""
    years = (CONDITION_ANCHOR - purchase_date).days / 365.25
    if years < 1:
        return "New"
    if years < 2:
        return "Good"
    if years < 3:
        return "Fair"
    return "Poor"


def read_source_rows(xlsx_path: Path) -> list[dict[str, str]]:
    """Read the spreadsheet exactly as given. No transformation beyond date formatting."""
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook["Assets"]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(cell).strip() for cell in rows[0]]
    expected = [
        "Asset Code",
        "Asset Name",
        "Category",
        "Employee Name",
        "Location",
        "Purchase Date",
    ]
    if header != expected:
        raise ValueError(f"Unexpected spreadsheet header: {header!r} (expected {expected!r})")

    records = []
    for raw in rows[1:]:
        if raw[0] is None:
            continue
        purchase = raw[5]
        if isinstance(purchase, datetime):
            purchase = purchase.date()
        records.append(
            {
                "asset_code": str(raw[0]).strip(),
                "asset_name": str(raw[1]).strip(),
                "category": str(raw[2]).strip(),
                "employee_name": str(raw[3]).strip(),
                "location": str(raw[4]).strip(),
                "purchase_date": purchase.isoformat(),
            }
        )
    workbook.close()
    return records


def _connect(db_path: Path) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def build(db_path: Path, xlsx_path: Path, reset: bool = True) -> dict[str, int]:
    source_rows = read_source_rows(xlsx_path)

    names_in_sheet = {row["employee_name"] for row in source_rows}
    missing = names_in_sheet - set(ORG_CHART)
    if missing:
        raise ValueError(
            f"Spreadsheet contains employees with no ORG_CHART entry: {sorted(missing)}. "
            "Add them to ORG_CHART in seed.py."
        )

    if reset and db_path.exists():
        db_path.unlink()

    conn = _connect(db_path)
    schema_sql = (Path(__file__).parent / "schema.sql").read_text()
    conn.executescript(schema_sql)

    now = datetime.now().isoformat(timespec="seconds")

    # --- employees (two passes so manager_id can reference a later row) ---
    employee_ids = {name: f"EMP{i:03d}" for i, name in enumerate(ORG_CHART, start=1)}
    for name, (department, location, _manager) in ORG_CHART.items():
        conn.execute(
            "INSERT INTO employees (employee_id, full_name, email, department, location,"
            " manager_id, source) VALUES (?, ?, ?, ?, ?, NULL, 'synthetic')",
            (employee_ids[name], name, _email_for(name), department, location),
        )
    for name, (_department, _location, manager) in ORG_CHART.items():
        if manager:
            conn.execute(
                "UPDATE employees SET manager_id = ? WHERE employee_id = ?",
                (employee_ids[manager], employee_ids[name]),
            )

    # --- assets from the spreadsheet, unchanged ---------------------------
    for row in source_rows:
        purchase = date.fromisoformat(row["purchase_date"])
        conn.execute(
            "INSERT INTO assets (asset_code, asset_name, category, employee_id, location,"
            " purchase_date, status, condition, source, updated_at)"
            " VALUES (?, ?, ?, ?, ?, ?, 'Assigned', ?, 'xlsx', ?)",
            (
                row["asset_code"],
                row["asset_name"],
                row["category"],
                employee_ids[row["employee_name"]],
                row["location"],
                row["purchase_date"],
                _condition_for(purchase),
                now,
            ),
        )

    # --- synthetic spare inventory ---------------------------------------
    for code, name, category, location, purchase_date, status, condition in SPARE_INVENTORY:
        conn.execute(
            "INSERT INTO assets (asset_code, asset_name, category, employee_id, location,"
            " purchase_date, status, condition, source, updated_at)"
            " VALUES (?, ?, ?, NULL, ?, ?, ?, ?, 'synthetic', ?)",
            (code, name, category, location, purchase_date, status, condition, now),
        )

    conn.commit()
    counts = {
        "employees": conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0],
        "assets_from_xlsx": conn.execute(
            "SELECT COUNT(*) FROM assets WHERE source = 'xlsx'"
        ).fetchone()[0],
        "assets_synthetic": conn.execute(
            "SELECT COUNT(*) FROM assets WHERE source = 'synthetic'"
        ).fetchone()[0],
    }
    conn.close()
    return counts


def verify(db_path: Path, xlsx_path: Path) -> list[str]:
    """Assert every source row still matches the spreadsheet. Returns list of problems."""
    problems: list[str] = []
    source_rows = {row["asset_code"]: row for row in read_source_rows(xlsx_path)}

    conn = _connect(db_path)
    db_rows = conn.execute(
        "SELECT a.asset_code, a.asset_name, a.category, a.location, a.purchase_date,"
        "       e.full_name AS employee_name"
        "  FROM assets a LEFT JOIN employees e ON e.employee_id = a.employee_id"
        " WHERE a.source = 'xlsx'"
    ).fetchall()
    conn.close()

    if len(db_rows) != len(source_rows):
        problems.append(
            f"row count mismatch: db has {len(db_rows)} xlsx-sourced rows, "
            f"spreadsheet has {len(source_rows)}"
        )

    for row in db_rows:
        expected = source_rows.get(row["asset_code"])
        if expected is None:
            problems.append(f"{row['asset_code']}: present in db but not in spreadsheet")
            continue
        for field in ("asset_name", "category", "location", "purchase_date", "employee_name"):
            if row[field] != expected[field]:
                problems.append(
                    f"{row['asset_code']}.{field}: db={row[field]!r} spreadsheet={expected[field]!r}"
                )
    return problems


def main(argv: list[str] | None = None) -> int:
    settings = get_settings()
    parser = argparse.ArgumentParser(description="Seed or verify the asset database.")
    parser.add_argument("--verify", action="store_true", help="check db against the spreadsheet")
    parser.add_argument("--db", type=Path, default=settings.db_path)
    parser.add_argument("--xlsx", type=Path, default=settings.source_xlsx)
    args = parser.parse_args(argv)

    if args.verify:
        problems = verify(args.db, args.xlsx)
        if problems:
            print("FAIL — database has drifted from the spreadsheet:")
            for problem in problems:
                print(f"  - {problem}")
            return 1
        print("OK — all spreadsheet-sourced rows match Sample_Asset_Master.xlsx")
        return 0

    counts = build(args.db, args.xlsx)
    print(f"Seeded {args.db}")
    print(f"  employees            {counts['employees']:>3}  (synthetic hierarchy)")
    print(f"  assets from xlsx     {counts['assets_from_xlsx']:>3}  (unchanged)")
    print(f"  assets synthetic     {counts['assets_synthetic']:>3}  (spare + in-repair stock)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
